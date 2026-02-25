#!/usr/bin/env python3
"""
Merge separate 'Extraheren resultaat ...xlsx' files into one workbook.

- Appends rows per sheet
- Skips header row for all files after the first one
- Continues on file errors and logs them
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook


DEFAULT_PREFIX = "Extraheren resultaat"
DEFAULT_SHEET_ORDER = [
    "Tekenhoofd",
    "Object",
    "Armgegevens",
    "Werkzaamheden",
    "as-Built",
    "Objecten dubbel",
]


def is_empty_row(values: Iterable[object]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and v.strip() == "":
            continue
        return False
    return True


def append_error(error_log: Path, source_file: Path, message: str) -> None:
    error_log.parent.mkdir(parents=True, exist_ok=True)
    new_file = not error_log.exists()
    with error_log.open("a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        if new_file:
            writer.writerow(["timestamp", "file", "error"])
        writer.writerow(
            [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                str(source_file),
                message.replace("\r", " ").replace("\n", " "),
            ]
        )


def find_source_files(source_dir: Path, prefix: str, output_path: Path) -> List[Path]:
    files = []
    for p in sorted(source_dir.glob("*.xlsx")):
        name = p.name.lower()
        if not name.startswith(prefix.lower()):
            continue
        if p.resolve() == output_path.resolve():
            continue
        if "samengevoegd" in name or "merged" in name:
            continue
        files.append(p)
    return files


def build_master_from_first(first_file: Path) -> Workbook:
    src_wb = load_workbook(first_file)
    master = Workbook()

    # Remove default sheet and recreate all source sheets in same order.
    if master.active:
        master.remove(master.active)

    for src_ws in src_wb.worksheets:
        dst_ws = master.create_sheet(src_ws.title)
        for row in src_ws.iter_rows(values_only=True):
            dst_ws.append(list(row))

    src_wb.close()
    return master


def ensure_sheet(master: Workbook, sheet_name: str):
    if sheet_name in master.sheetnames:
        return master[sheet_name]
    return master.create_sheet(sheet_name)


def append_workbook(master: Workbook, source_file: Path, skip_header: bool = True) -> None:
    src_wb = load_workbook(source_file, data_only=False)
    try:
        for src_ws in src_wb.worksheets:
            dst_ws = ensure_sheet(master, src_ws.title)
            start_row_idx = 1 if skip_header else 0  # 0-based enumerate over rows
            for idx, row in enumerate(src_ws.iter_rows(values_only=True)):
                if idx < start_row_idx:
                    continue
                if is_empty_row(row):
                    continue
                dst_ws.append(list(row))
    finally:
        src_wb.close()


def reorder_sheets(master: Workbook) -> None:
    ordered = []
    for name in DEFAULT_SHEET_ORDER:
        if name in master.sheetnames:
            ordered.append(master[name])
    for ws in master.worksheets:
        if ws.title not in DEFAULT_SHEET_ORDER:
            ordered.append(ws)
    master._sheets = ordered  # acceptable internal use for ordering


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge Java naverwerking output workbooks into one workbook.")
    parser.add_argument("--source-dir", default="Doel", help="Folder with 'Extraheren resultaat ...xlsx' files")
    parser.add_argument(
        "--output",
        default="Eindresultaat/Extraheren resultaat SAMENGEVOEGD.xlsx",
        help="Output merged workbook path",
    )
    parser.add_argument(
        "--prefix",
        default=DEFAULT_PREFIX,
        help="Filename prefix to include (default: 'Extraheren resultaat')",
    )
    parser.add_argument(
        "--error-log",
        default="Eindresultaat/merge_errors.csv",
        help="CSV log for merge errors (files that are skipped)",
    )
    args = parser.parse_args()

    source_dir = Path(args.source_dir)
    output_path = Path(args.output)
    error_log = Path(args.error_log)

    if not source_dir.exists():
        print(f"Source dir not found: {source_dir}")
        return 2

    files = find_source_files(source_dir, args.prefix, output_path)
    if not files:
        print("No source result files found to merge.")
        return 1

    print(f"Found {len(files)} result files to merge.")

    master = None
    success = 0
    failed = 0

    for idx, f in enumerate(files):
        try:
            if master is None:
                master = build_master_from_first(f)
            else:
                append_workbook(master, f, skip_header=True)
            success += 1
            print(f"[OK] {idx+1}/{len(files)} {f.name}")
        except Exception as exc:
            failed += 1
            append_error(error_log, f, f"{exc.__class__.__name__}: {exc}")
            print(f"[ERR] {idx+1}/{len(files)} {f.name}: {exc}")
            continue

    if master is None:
        print("No valid workbook could be merged.")
        return 1

    reorder_sheets(master)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    master.save(output_path)
    master.close()

    print("")
    print(f"Merged output: {output_path}")
    print(f"Success: {success}, Failed: {failed}")
    if failed:
        print(f"Merge errors: {error_log}")

    return 0 if success else 1


if __name__ == "__main__":
    raise SystemExit(main())
